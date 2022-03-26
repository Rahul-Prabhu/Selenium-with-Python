from selenium import webdriver
from selenium.webdriver import ActionChains
import time
from selenium.webdriver.chrome.options import Options
import openpyxl

path="C:\SeleniumProject\Train.xlsx"

workbook=openpyxl.load_workbook(path)

sheet=workbook.active

rows=sheet.max_row
cols=sheet.max_column

print(rows)
print(cols)

for r in range(1,rows+1):
    for c in range(1,cols+1):
        print(sheet.cell(row=r,column=c).value,"|",end="            ")
    print()